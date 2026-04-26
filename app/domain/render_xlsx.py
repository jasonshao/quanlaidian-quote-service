#!/usr/bin/env python3
"""
XLSX quotation generator — ported from generate_quotation.py (XLSX portion).

Changes from original:
1. Returns bytes from BytesIO instead of writing to file
2. Shared helpers imported from render_pdf.py
3. main() / argparse removed
4. Tiered sheet addition controlled by render_xlsx() entry point
"""

from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path

from app.domain.render_pdf import (
    fmt_money,
    get_deal_price_factor,
    get_item_unit_price,
    get_item_subtotal,
    get_item_cost_unit_price,
    get_item_cost_subtotal,
    get_tier_unit_price,
    number_to_chinese,
    gen_quote_number,
    calc_actual_price,
    fmt_pct,
    _LOGO_SHOUQIANBA,
    _LOGO_QUANLAIDIAN,
)


# ============================================================
# 水印配置（与 PDF 保持一致）
# ============================================================
_WATERMARK_TEXT_COLOR = (170, 170, 170)   # RGB，浅灰 #AAAAAA，与 PDF 一致
_WATERMARK_ALPHA = 0.22                    # 透明度 0=透明，1=不透明
_WATERMARK_FONT_SIZE = 24                 # 水印字号（像素，2000×2000 画布）
_WATERMARK_ANGLE = -30                    # 倾斜角度（度）
_WATERMARK_TILE_X_SPACING = 450           # 水印 x 方向重复间距（像素）
_WATERMARK_TILE_Y_SPACING = 280           # 水印 y 方向重复间距（像素）


def _generate_watermark_image(quote_no: str, quote_date: str) -> bytes:
    """生成半透明斜向平铺水印 PNG 图片（用于 Excel 背景）。

    水印策略：
    - PIL 绘制，中灰 (136,136,136)，透明度 0.30
    - 先在未旋转画布上按网格平铺文字，再整体旋转 -30°
    - 返回 PNG 字节流，供 openpyxl 嵌入为浮动图像
    """
    from PIL import Image as PILImage, ImageDraw as PILImageDraw, ImageFont as PILImageFont
    from io import BytesIO

    text = f"{quote_no}  {quote_date}".strip()
    if not text:
        return b''

    img_w, img_h = 2000, 2000
    canvas = PILImage.new('RGBA', (img_w, img_h), (255, 255, 255, 0))
    draw = PILImageDraw.Draw(canvas)

    # CJK 字体候选（Linux/macOS 常见位置）
    font = None
    candidates = [
        '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
    ]
    for path in candidates:
        from pathlib import Path
        if Path(path).exists():
            try:
                font = PILImageFont.truetype(path, _WATERMARK_FONT_SIZE)
                break
            except Exception:
                continue

    if font is None:
        try:
            font = PILImageFont.load_default()
        except Exception:
            return b''

    alpha = max(1, min(255, int(255 * _WATERMARK_ALPHA)))
    text_color = _WATERMARK_TEXT_COLOR + (alpha,)

    # 简单均匀网格平铺（未旋转画布上）
    step_x = _WATERMARK_TILE_X_SPACING
    step_y = _WATERMARK_TILE_Y_SPACING
    y = 0
    row_idx = 0
    while y < img_h:
        # 奇偶行半步错开
        x_offset = (step_x // 2) if (row_idx % 2) else 0
        x = -x_offset
        while x < img_w:
            draw.text((x, y), text, font=font, fill=text_color)
            x += step_x
        y += step_y
        row_idx += 1

    # 整体旋转 -30°
    rotated = canvas.rotate(_WATERMARK_ANGLE, expand=True, fillcolor=(255, 255, 255, 0))

    buf = BytesIO()
    rotated.save(buf, format='PNG')
    return buf.getvalue()


# ============================================================
# Header logos — inserted at top-left of every sheet's row 1.
# Caller MUST leave row 1 free (content starts from row 2).
# ============================================================
_XL_LOGO_HEIGHT_PX = 26       # ~7mm at 96 DPI, matches PDF logo size
_XL_LOGO_GAP_PX = 12           # ~4mm horizontal gap between the two logos
_XL_LOGO_ROW_HEIGHT_PT = 26    # ~35px row height, leaves ~9px padding
_XL_LOGO_LEFT_OFFSET_PX = 4    # small left inset inside column A


def _xl_add_header_logos(ws):
    """Anchor 收钱吧 + 全来店 logos side-by-side in the top-left of `ws`.

    Row 1's height is set to reserve space for the logos. The caller is
    responsible for ensuring no other content lives in row 1.

    Uses AbsoluteAnchor (pixel coordinates relative to the sheet's top-left)
    so placement is independent of column A's width — a OneCellAnchor with
    colOff=102px on a sheet where column A is only ~49px wide caused Numbers
    to collapse both logos onto the same spot on every sheet except the
    tier-reference sheet (which happens to have a wide column A).

    Logos are pre-resized via PIL before embedding so renderers that ignore
    the anchor `ext` (e.g. Numbers, Preview) still display them at the right
    size — the source PNG for 收钱吧 is 2123×641 native, so leaving it
    unscaled would cause it to cover the second logo.
    """
    from io import BytesIO

    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
    from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    ws.row_dimensions[1].height = _XL_LOGO_ROW_HEIGHT_PT

    x_cursor_px = _XL_LOGO_LEFT_OFFSET_PX
    for path in (_LOGO_SHOUQIANBA, _LOGO_QUANLAIDIAN):
        if not path.exists():
            continue

        scaled_w = None
        try:
            with PILImage.open(path) as pil:
                src_w, src_h = pil.size
                aspect = src_w / src_h if src_h else 1.0
                scaled_h = _XL_LOGO_HEIGHT_PX
                scaled_w = max(1, int(round(scaled_h * aspect)))
                # Downsample the PNG itself so naive viewers (which ignore
                # the anchor ext) still render at the intended display size.
                resized = pil.convert("RGBA").resize((scaled_w, scaled_h), PILImage.LANCZOS)
                buf = BytesIO()
                resized.save(buf, format="PNG")
                buf.seek(0)

            img = XLImage(buf)
            img.width = scaled_w
            img.height = scaled_h

            pos = XDRPoint2D(
                x=pixels_to_EMU(x_cursor_px),
                y=pixels_to_EMU(2),
            )
            ext = XDRPositiveSize2D(
                cx=pixels_to_EMU(scaled_w),
                cy=pixels_to_EMU(scaled_h),
            )
            img.anchor = AbsoluteAnchor(pos=pos, ext=ext)
            ws.add_image(img)
        except Exception:
            scaled_w = None

        if scaled_w is not None:
            x_cursor_px += scaled_w + _XL_LOGO_GAP_PX


def _xl_header_style(cell):
    """Excel 表头单元格样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', fgColor='FFB300')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _xl_title_style(cell, size=16):
    """Excel 标题单元格样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, size=size, color='CC8800')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _xl_subtitle_style(cell):
    """Excel 副标题单元格样式"""
    from openpyxl.styles import Font, Alignment
    cell.font = Font(name='微软雅黑', size=11, color='555555')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def _xl_info_label_style(cell):
    """Excel 客户信息标签样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
    cell.fill = PatternFill('solid', fgColor='FFFBF0')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def _xl_info_value_style(cell):
    """Excel 客户信息值样式"""
    from openpyxl.styles import Font, Alignment
    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def _xl_data_style(cell, align='center', bold=False, bg=None, num_format=None):
    """Excel 数据单元格样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if num_format:
        cell.number_format = num_format

def _xl_total_style(cell, align='right'):
    """Excel 合计行样式"""
    from openpyxl.styles import Font, PatternFill, Alignment
    cell.font = Font(name='微软雅黑', bold=True, size=10)
    cell.fill = PatternFill('solid', fgColor='FFF5D6')
    cell.alignment = Alignment(horizontal=align, vertical='center')

def _xl_apply_border(ws, min_row, min_col, max_row, max_col):
    """为 Excel 单元格区域应用细边框"""
    from openpyxl.styles import Border, Side
    thin = Side(style='thin', color='D0D5DD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, min_col=min_col,
                             max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border

def _xl_write_item_table(ws, items, start_row, sheet_name='', compute_values=False):
    """
    向 worksheet 写入报价明细表。
    列：A=序号, B=商品分类, C=商品名称, D=单位, E=数量, F=商品单价, G=小计, H=功能说明
    compute_values=True 时直接写计算后的数值（兼容性更好），
    否则写 Excel 公式（便于手动调整）。
    返回: (最后数据行号, 合计行号)
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    # 表头
    headers = ['序号', '商品分类', '商品名称', '单位', '数量', '商品单价', '小计', '功能说明']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        _xl_header_style(cell)
    ws.row_dimensions[start_row].height = 22

    data_start = start_row + 1
    current_row = data_start

    for idx, item in enumerate(items, 1):
        qty = item.get('数量', 1)
        unit_price_d = get_item_unit_price(item)
        subtotal_d = get_item_subtotal(item)
        is_gift = (unit_price_d == '赠送' or subtotal_d == '赠送')

        # 交替行背景
        bg = 'F5F7FA' if idx % 2 == 0 else None

        # A: 序号
        c = ws.cell(row=current_row, column=1, value=idx)
        _xl_data_style(c, align='center', bg=bg)

        # B: 商品分类
        c = ws.cell(row=current_row, column=2, value=item.get('商品分类', ''))
        _xl_data_style(c, align='left', bg=bg)

        # C: 商品名称
        c = ws.cell(row=current_row, column=3, value=item.get('商品名称', ''))
        _xl_data_style(c, align='left', bg=bg)

        # D: 单位
        c = ws.cell(row=current_row, column=4, value=item.get('单位', ''))
        _xl_data_style(c, align='center', bg=bg)

        # E: 数量
        c = ws.cell(row=current_row, column=5, value=int(qty))
        _xl_data_style(c, align='center', bg=bg)

        # F: 商品单价
        if is_gift:
            c = ws.cell(row=current_row, column=6, value='赠送')
            _xl_data_style(c, align='center', bg=bg)
        else:
            c = ws.cell(row=current_row, column=6, value=int(unit_price_d))
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0')

        # G: 小计
        if is_gift:
            c = ws.cell(row=current_row, column=7, value='赠送')
            _xl_data_style(c, align='center', bg=bg)
        else:
            c = ws.cell(row=current_row, column=7, value=int(subtotal_d))
            _xl_data_style(c, align='right', bg=bg, num_format='#,##0')

        # H: 功能说明 (multi-line wrap)
        description = item.get('功能说明', '') or ''
        c = ws.cell(row=current_row, column=8, value=description)
        _xl_data_style(c, align='left', bg=bg)

        # Row height scales with description line count so wrapped text stays
        # readable. Packages often have 4–6 bulleted lines; simple modules one.
        line_count = max(1, description.count('\n') + 1) if description else 1
        ws.row_dimensions[current_row].height = max(18, 14 * min(line_count, 6))
        current_row += 1

        # ── 子行：门店套餐展开为「商户中心 / 收银系统 / ...」子模块 ──
        sub_bg = 'F9FAFB'
        for sub in (item.get('子项') or []):
            # A: 序号留空
            c = ws.cell(row=current_row, column=1, value='')
            _xl_data_style(c, align='center', bg=sub_bg)
            # B: 商品分类
            c = ws.cell(row=current_row, column=2, value=sub.get('商品分类', ''))
            _xl_data_style(c, align='left', bg=sub_bg)
            # C: 商品名称（以两个全角空格缩进以显示层级）
            c = ws.cell(row=current_row, column=3, value=f"　　{sub.get('商品名称', '')}")
            _xl_data_style(c, align='left', bg=sub_bg)
            # D: 单位
            c = ws.cell(row=current_row, column=4, value=sub.get('单位', ''))
            _xl_data_style(c, align='center', bg=sub_bg)
            # E/F/G: 数量 / 单价 / 小计 均显示 "-"
            for col in (5, 6, 7):
                c = ws.cell(row=current_row, column=col, value='-')
                _xl_data_style(c, align='center', bg=sub_bg)
            # H: 子模块功能说明
            sub_desc = sub.get('功能说明', '') or ''
            c = ws.cell(row=current_row, column=8, value=sub_desc)
            _xl_data_style(c, align='left', bg=sub_bg)
            sub_lines = max(1, sub_desc.count('\n') + 1) if sub_desc else 1
            ws.row_dimensions[current_row].height = max(16, 13 * min(sub_lines, 5))
            current_row += 1

    last_data_row = current_row - 1

    # 合计行
    total_row = current_row
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=5)
    c = ws.cell(row=total_row, column=1, value='合计')
    _xl_total_style(c, align='center')

    ws.cell(row=total_row, column=6, value='')
    _xl_total_style(ws.cell(row=total_row, column=6))

    # 合计公式：SUM忽略文本（赠送）
    c = ws.cell(row=total_row, column=7,
                value=f'=SUM(G{data_start}:G{last_data_row})')
    _xl_total_style(c, align='right')
    c.number_format = '#,##0'

    # H: 功能说明 合计行留空
    _xl_total_style(ws.cell(row=total_row, column=8))

    ws.row_dimensions[total_row].height = 20

    # 应用边框（含表头）
    _xl_apply_border(ws, start_row, 1, total_row, 8)

    return last_data_row, total_row


# 8 列标准宽度（明细表基准），各 sheet 视觉宽度以此为准统一。
_XL_STANDARD_COL_WIDTHS = {
    'A': 7,   # 序号
    'B': 16,  # 商品分类
    'C': 28,  # 商品名称
    'D': 10,  # 单位
    'E': 8,   # 数量
    'F': 14,  # 商品单价
    'G': 16,  # 小计
    'H': 42,  # 功能说明
}
_XL_STANDARD_TOTAL_WIDTH = sum(_XL_STANDARD_COL_WIDTHS.values())  # 141


def _xl_set_col_widths(ws):
    """设置标准列宽"""
    for col, width in _XL_STANDARD_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _xl_write_annotation_block(ws, annotation, start_row, end_col_letter='H'):
    """Write a merged annotation block below the item table.

    Layout:
      row N:   [title]      —  golden accent header
      row N+1: [category]   —  subtitle
      row N+2: * text_line1 —  grey body
      row N+3: * text_line2
      ...
    Returns the next row after the block.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    end_col = {c: i for i, c in enumerate('ABCDEFGHIJKLMN', 1)}[end_col_letter]
    r = start_row

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
    c = ws.cell(row=r, column=1, value=annotation.get('title', ''))
    c.font = Font(name='微软雅黑', bold=True, size=11, color='CC8800')
    c.fill = PatternFill('solid', fgColor='FFF5D6')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 22
    r += 1

    if annotation.get('category'):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
        c = ws.cell(row=r, column=1, value=annotation['category'])
        c.font = Font(name='微软雅黑', bold=True, size=10, color='555555')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 18
        r += 1

    for line in annotation.get('text_lines') or []:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
        c = ws.cell(row=r, column=1, value=f'* {line}')
        c.font = Font(name='微软雅黑', size=9, color='555555')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 32
        r += 1

    return r


# ============================================================
# 标准模板（≤50店）— Excel
# ============================================================
def _generate_xlsx_standard(data):
    """生成标准单页 Excel 报价单（≤50店）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '报价单'
    ws.sheet_view.showGridLines = False

    _xl_set_col_widths(ws)

    # ── Logo 行（行1）──
    _xl_add_header_logos(ws)

    # ── 标题区（行2-3） ──
    ws.merge_cells('A2:H2')
    c = ws.cell(row=2, column=1, value='"全来店"产品报价单')
    _xl_title_style(c, size=16)
    ws.row_dimensions[2].height = 36

    ws.merge_cells('A3:H3')
    c = ws.cell(row=3, column=1, value='上海收钱吧互联网科技股份有限公司')
    _xl_subtitle_style(c)
    ws.row_dimensions[3].height = 24

    # ── 空行 ──
    ws.row_dimensions[4].height = 6

    # ── 客户信息区（行5-8） ──
    client = data.get('客户信息', {})
    quote_no = data.get('报价编号', gen_quote_number())
    quote_date = data.get('报价日期', datetime.now().strftime('%Y年%m月%d日'))
    validity = data.get('报价有效期', '30个工作日')

    info_rows = [
        (f'致：{client.get("公司名称", "")}',   f'报价编号：{quote_no}'),
        (f'联系人：{client.get("联系人", "")}',  f'报价日期：{quote_date}'),
        (f'地址：{client.get("地址", "")}',       f'有效期：{validity}'),
        (f'电话：{client.get("电话", "")}',       ''),
    ]

    for i, (left, right) in enumerate(info_rows):
        row_num = 5 + i
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=4)
        c = ws.cell(row=row_num, column=1, value=left)
        _xl_info_value_style(c)

        ws.merge_cells(start_row=row_num, start_column=5,
                       end_row=row_num, end_column=8)
        c = ws.cell(row=row_num, column=5, value=right)
        _xl_info_value_style(c)
        ws.row_dimensions[row_num].height = 18

    # ── 空行 ──
    ws.row_dimensions[9].height = 8

    # ── 报价明细表（从行10开始） ──
    items = data.get('报价项目', [])
    header_row = 10
    _, total_row = _xl_write_item_table(ws, items, header_row)

    # ── 金额大写（合计行下方） ──
    notes_row = total_row + 2
    ws.merge_cells(start_row=notes_row, start_column=1,
                   end_row=notes_row, end_column=8)
    total_val = Decimal('0')
    for item in items:
        subtotal = get_item_subtotal(item)
        if subtotal != '赠送':
            total_val += subtotal

    chinese_amt = number_to_chinese(float(total_val))
    c = ws.cell(row=notes_row, column=1,
                value=f'合计金额（大写）：{chinese_amt}')
    c.font = Font(name='微软雅黑', size=10, bold=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[notes_row].height = 20

    # ── 权益类自助充值模块 注释区块 ──
    annotation_cursor = notes_row + 1
    for annotation in (data.get('附加说明') or []):
        if not annotation:
            continue
        annotation_cursor = _xl_write_annotation_block(
            ws, annotation, annotation_cursor, end_col_letter='H',
        )

    # ── 备注条款 ──
    terms = data.get('条款', [
        '以上报价金额均为含税金额，税率为6%；',
        '报价有效期为30个工作日，自报价单生成之日起；',
        '具体折扣金额按签订合同（或销售订单）时具体数量确定价格；',
        '涉及短信、小程序授权、外卖平台接口调用等第三方机构收费部分，需单独计费；',
        '如需要三方代仓对接，需要一事一议。',
    ])

    terms_start = annotation_cursor
    ws.merge_cells(start_row=terms_start, start_column=1,
                   end_row=terms_start, end_column=8)
    c = ws.cell(row=terms_start, column=1, value='备注：')
    c.font = Font(name='微软雅黑', size=10, bold=True, color='CC8800')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[terms_start].height = 18

    cn_nums = '①②③④⑤⑥⑦⑧⑨⑩'
    for i, term in enumerate(terms):
        r = terms_start + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        prefix = cn_nums[i] if i < 10 else f'{i+1}.'
        c = ws.cell(row=r, column=1, value=f'{prefix} {term}')
        c.font = Font(name='微软雅黑', size=9, color='555555')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 16

    # ── 页脚 ──
    footer_row = terms_start + 1 + len(terms) + 1
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=8)
    c = ws.cell(row=footer_row, column=1,
                value='上海收钱吧互联网科技股份有限公司  |  地址：上海市闵行区浦江智慧广场陈行公路2168号7号楼')
    c.font = Font(name='微软雅黑', size=8, color='888888')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[footer_row].height = 16

    # 冻结表头
    ws.freeze_panes = f'A{header_row + 1}'

    return wb


# ============================================================
# 定制多页模板（>50店）— Excel
# ============================================================
def _generate_xlsx_custom(data):
    """生成定制多 Sheet Excel 报价单（>50店）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    client = data.get('客户信息', {})
    quote_no = data.get('报价编号', gen_quote_number())
    quote_date = data.get('报价日期', datetime.now().strftime('%Y-%m-%d'))
    validity = data.get('报价有效期', '30个工作日')
    items = data.get('报价项目', [])

    # ── Sheet 1：封面 ──
    ws_cover = wb.create_sheet('封面')
    ws_cover.sheet_view.showGridLines = False
    # 与明细表 sheet 统一列宽（8 列 共 141 单位）
    _xl_set_col_widths(ws_cover)

    _xl_add_header_logos(ws_cover)
    ws_cover.merge_cells('A2:H2')
    c = ws_cover.cell(row=2, column=1, value='"全来店"产品报价方案')
    _xl_title_style(c, size=18)
    ws_cover.row_dimensions[2].height = 44

    ws_cover.merge_cells('A3:H3')
    c = ws_cover.cell(row=3, column=1, value='上海收钱吧互联网科技股份有限公司')
    _xl_subtitle_style(c)
    ws_cover.row_dimensions[3].height = 28

    ws_cover.row_dimensions[4].height = 12

    cover_info = [
        ('客户名称', client.get('公司名称', '')),
        ('联系人',   client.get('联系人', '')),
        ('联系电话', client.get('电话', '')),
        ('报价编号', quote_no),
        ('报价日期', quote_date),
        ('有效期',   validity),
    ]

    # 标签占 A:B 两列（23 单位），值跨 C:H 六列（118 单位），整体与明细表等宽。
    for i, (label, value) in enumerate(cover_info):
        r = 5 + i
        ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c_label = ws_cover.cell(row=r, column=1, value=label)
        _xl_info_label_style(c_label)
        ws_cover.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        c_val = ws_cover.cell(row=r, column=3, value=value)
        _xl_info_value_style(c_val)
        ws_cover.row_dimensions[r].height = 22

    _xl_apply_border(ws_cover, 5, 1, 5 + len(cover_info) - 1, 8)

    # 记录封面下一可用行（用于后续追加汇总内容）
    cover_summary_start_row = 5 + len(cover_info) + 2

    # ── 按模块分类（不含硬件设备）──
    categories = {
        '门店软件套餐': [],
        '门店增值模块': [],
        '总部模块': [],
        '实施服务': [],
    }

    for item in items:
        cat = item.get('模块分类', '门店软件套餐')
        if cat in categories:
            categories[cat].append(item)
        elif cat != '硬件设备':
            categories['门店软件套餐'].append(item)

    # ── 各分类 Sheet ──
    def _override_items(src_items, deal_price_factor, qty=1):
        """返回 qty/商品单价 覆盖后的副本（刊例价展示用）"""
        result = []
        for it in src_items:
            ni = dict(it)
            ni['数量'] = qty
            if ni.get('标准价') not in ('赠送', None):
                if ni.get('模块分类') == '门店软件套餐':
                    unit_price = calc_actual_price(ni.get('标准价'), deal_price_factor)
                    ni['deal_price_factor'] = deal_price_factor
                    ni['成交价系数'] = deal_price_factor
                    ni['折扣'] = round(1 - deal_price_factor, 6)
                    ni['商品单价'] = int(unit_price)
                subtotal = get_item_unit_price(ni)
                if subtotal != '赠送':
                    ni['报价小计'] = int((subtotal * Decimal(str(qty))).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
            result.append(ni)
        return result

    MERGED_SHEET_CATS = ['门店软件套餐', '门店增值模块']
    cat_totals = {}

    # ── 合并 Sheet：门店软件与增值 ──
    merged_has_items = any(categories[c] for c in MERGED_SHEET_CATS)
    if merged_has_items:
        ws_merged = wb.create_sheet('门店软件与增值')
        ws_merged.sheet_view.showGridLines = False
        _xl_set_col_widths(ws_merged)

        _xl_add_header_logos(ws_merged)
        ws_merged.merge_cells('A2:H2')
        c = ws_merged.cell(row=2, column=1, value='门店软件与增值模块')
        _xl_title_style(c, size=14)
        ws_merged.row_dimensions[2].height = 30
        ws_merged.row_dimensions[3].height = 8

        section_row = 4
        for cat_name in MERGED_SHEET_CATS:
            cat_items = categories[cat_name]
            if not cat_items:
                continue
            # 分区标题行
            ws_merged.merge_cells(start_row=section_row, start_column=1,
                                  end_row=section_row, end_column=8)
            c = ws_merged.cell(row=section_row, column=1, value=cat_name)
            c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
            c.fill = PatternFill('solid', fgColor='FFFBF0')
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws_merged.row_dimensions[section_row].height = 18
            section_row += 1

            display_items = _override_items(cat_items, deal_price_factor=0.8, qty=1)
            _, total_row = _xl_write_item_table(ws_merged, display_items, section_row, compute_values=True)
            cat_totals[cat_name] = (f'G{total_row}', ws_merged.title)
            section_row = total_row + 2

        ws_merged.freeze_panes = 'A5'

    # ── 总部模块 Sheet ──
    if categories.get('总部模块'):
        ws = wb.create_sheet('总部模块')
        ws.sheet_view.showGridLines = False
        _xl_set_col_widths(ws)
        _xl_add_header_logos(ws)
        ws.merge_cells('A2:H2')
        c = ws.cell(row=2, column=1, value='总部模块')
        _xl_title_style(c, size=14)
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 8
        display_items = _override_items(categories['总部模块'], deal_price_factor=0.8, qty=1)
        _, total_row = _xl_write_item_table(ws, display_items, 4, compute_values=True)
        ws.freeze_panes = 'A5'
        cat_totals['总部模块'] = (f'G{total_row}', ws.title)

    # ── 实施服务 Sheet ──
    if categories.get('实施服务'):
        ws = wb.create_sheet('实施服务')
        ws.sheet_view.showGridLines = False
        _xl_set_col_widths(ws)
        _xl_add_header_logos(ws)
        ws.merge_cells('A2:H2')
        c = ws.cell(row=2, column=1, value='实施服务')
        _xl_title_style(c, size=14)
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 8
        display_items = _override_items(categories['实施服务'], deal_price_factor=1.0, qty=1)
        _, total_row = _xl_write_item_table(ws, display_items, 4, compute_values=True)
        ws.freeze_panes = 'A5'
        cat_totals['实施服务'] = (f'G{total_row}', ws.title)

    # ── 封面追加：权益类自助充值模块 注释区块 ──
    r = cover_summary_start_row
    for annotation in (data.get('附加说明') or []):
        if not annotation:
            continue
        r = _xl_write_annotation_block(ws_cover, annotation, r, end_col_letter='H')

    # ── 封面追加：条款说明 ──
    terms = data.get('条款', [
        '以上报价金额均为含税金额，税率为6%；',
        '报价有效期为30个工作日，自报价单生成之日起；',
        '具体折扣金额按签订合同（或销售订单）时具体数量确定价格；',
        '涉及短信、小程序授权、外卖平台接口调用等第三方机构收费部分，需单独计费；',
        '如需要三方代仓对接，需要一事一议。',
    ])
    ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws_cover.cell(row=r, column=1, value='备注与条款：')
    c.font = Font(name='微软雅黑', size=10, bold=True, color='CC8800')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws_cover.row_dimensions[r].height = 18
    r += 1

    cn_nums = '①②③④⑤⑥⑦⑧⑨⑩'
    for i, term in enumerate(terms):
        ws_cover.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        prefix = cn_nums[i] if i < 10 else f'{i+1}.'
        c = ws_cover.cell(row=r, column=1, value=f'{prefix} {term}')
        c.font = Font(name='微软雅黑', size=9, color='555555')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws_cover.row_dimensions[r].height = 16
        r += 1

    # NOTE: tiered sheet is added by render_xlsx() entry point, not here

    return wb


# ============================================================
# 阶梯报价参考 Sheet
# ============================================================
def _xl_add_tiered_sheet(wb, data):
    """在 Excel 工作簿末尾添加阶梯报价参考 Sheet（9 列：序号/分类/名称/单位/锚点1单价/锚点1小计/锚点2单价/锚点2小计/说明）"""
    tiers = data.get('阶梯配置', [])
    # 规格要求固定 2 档（锚点1/锚点2）。不足 2 档不渲染。
    if len(tiers) < 2:
        return

    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet('阶梯报价参考')
    ws.sheet_view.showGridLines = False

    items = data.get('报价项目', [])
    tier_low, tier_high = tiers[0], tiers[1]

    # 列宽：总和对齐 _XL_STANDARD_TOTAL_WIDTH=141（明细表基准），I 列吸收余量
    widths = {'A': 7, 'B': 16, 'C': 26, 'D': 10, 'E': 14, 'F': 16, 'G': 14, 'H': 16, 'I': 22}
    diff = _XL_STANDARD_TOTAL_WIDTH - sum(widths.values())
    if diff:
        widths['I'] += diff
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Logo 行（行1）
    _xl_add_header_logos(ws)

    # 标题
    ws.merge_cells('A2:I2')
    c = ws.cell(row=2, column=1, value='阶梯报价参考')
    _xl_title_style(c, size=14)
    ws.row_dimensions[2].height = 32

    headers = [
        '序号', '商品分类', '商品名称', '单位',
        f"{tier_low['标签']} 单价", f"{tier_low['标签']} 小计",
        f"{tier_high['标签']} 单价", f"{tier_high['标签']} 小计",
        '功能说明',
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        _xl_header_style(c)
    ws.row_dimensions[3].height = 22

    current_row = 4
    seq = 0

    cat_order = ['门店软件套餐', '门店增值模块', '总部模块', '实施服务']
    categories = {k: [] for k in cat_order}
    for item in items:
        cat = item.get('模块分类', '门店软件套餐')
        if cat == '硬件设备':
            continue
        if cat in categories:
            categories[cat].append(item)
        else:
            categories['门店软件套餐'].append(item)

    grand_totals = [Decimal('0'), Decimal('0')]
    NO_TIER_DISCOUNT_CATS = {'实施服务'}

    SUBTOTAL_BG = 'FFF5D6'
    TOTAL_BG = 'FFE082'

    def _apply_bg(cells, bg):
        for cell in cells:
            cell.fill = PatternFill('solid', fgColor=bg)

    for cat_name in cat_order:
        cat_items = categories[cat_name]
        if not cat_items:
            continue

        # 分类标题行：跨 A-I 合并
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=9)
        c = ws.cell(row=current_row, column=1, value=cat_name)
        c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
        c.fill = PatternFill('solid', fgColor='FFFBF0')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        cat_subtotals = [Decimal('0'), Decimal('0')]
        apply_tier_discount = cat_name not in NO_TIER_DISCOUNT_CATS

        for item in cat_items:
            seq += 1
            unit = item.get('单位', '')
            item_qty = item.get('数量', 1)
            is_per_store = '店' in unit
            unit_price_item = get_item_unit_price(item)
            is_gift = unit_price_item == '赠送'

            # A: 序号
            c = ws.cell(row=current_row, column=1, value=seq)
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='center', vertical='center')
            # B: 商品分类
            c = ws.cell(row=current_row, column=2, value=item.get('商品分类', ''))
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='left', vertical='center')
            # C: 商品名称
            c = ws.cell(row=current_row, column=3, value=item.get('商品名称', ''))
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='left', vertical='center')
            # D: 单位
            c = ws.cell(row=current_row, column=4, value=unit)
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='center', vertical='center')

            for ti, tier in enumerate((tier_low, tier_high)):
                col_unit = 5 + ti * 2  # E=5, G=7
                col_sub = 6 + ti * 2   # F=6, H=8
                if is_gift:
                    for col in (col_unit, col_sub):
                        cc = ws.cell(row=current_row, column=col, value='赠送')
                        cc.font = Font(name='微软雅黑', size=9)
                        cc.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    d = Decimal(str(get_deal_price_factor(tier))) if apply_tier_discount else Decimal('1')
                    actual_price = get_tier_unit_price(item, d) if apply_tier_discount else unit_price_item
                    qty = Decimal(str(tier['门店数'])) if is_per_store else Decimal(str(item_qty))
                    subtotal = (actual_price * qty).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
                    cat_subtotals[ti] += subtotal
                    # 单价
                    cu = ws.cell(row=current_row, column=col_unit, value=int(actual_price))
                    cu.font = Font(name='微软雅黑', size=9)
                    cu.number_format = '#,##0'
                    cu.alignment = Alignment(horizontal='right', vertical='center')
                    # 小计
                    cs = ws.cell(row=current_row, column=col_sub, value=int(subtotal))
                    cs.font = Font(name='微软雅黑', size=9)
                    cs.number_format = '#,##0'
                    cs.alignment = Alignment(horizontal='right', vertical='center')

            # I: 功能说明
            description = item.get('功能说明', '') or ''
            c = ws.cell(row=current_row, column=9, value=description)
            c.font = Font(name='微软雅黑', size=9)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            line_count = max(1, description.count('\n') + 1) if description else 1
            ws.row_dimensions[current_row].height = max(18, 13 * min(line_count, 6))
            current_row += 1

        # 分类小计行
        c = ws.cell(row=current_row, column=2, value='小计')
        c.font = Font(name='微软雅黑', bold=True, size=9)
        c.alignment = Alignment(horizontal='center', vertical='center')
        for ti, sub in enumerate(cat_subtotals):
            grand_totals[ti] += sub
            col = 6 + ti * 2  # F=6, H=8
            cc = ws.cell(row=current_row, column=col, value=int(sub))
            cc.font = Font(name='微软雅黑', bold=True, size=9)
            cc.number_format = '#,##0'
            cc.alignment = Alignment(horizontal='right', vertical='center')
        _apply_bg((ws.cell(row=current_row, column=col) for col in range(1, 10)), SUBTOTAL_BG)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # 合计行
    c = ws.cell(row=current_row, column=2, value='合计')
    c.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ti, tot in enumerate(grand_totals):
        col = 6 + ti * 2
        cc = ws.cell(row=current_row, column=col, value=int(tot))
        cc.font = Font(name='微软雅黑', bold=True, size=10, color='CC8800')
        cc.number_format = '#,##0'
        cc.alignment = Alignment(horizontal='right', vertical='center')
    _apply_bg((ws.cell(row=current_row, column=col) for col in range(1, 10)), TOTAL_BG)
    ws.row_dimensions[current_row].height = 22

    _xl_apply_border(ws, 3, 1, current_row, 9)


# ============================================================
# 公共入口
# ============================================================
def render_xlsx(config: dict) -> bytes:
    """Generate XLSX quotation from config dict. Returns XLSX bytes.

    Watermark is injected as a page-header image (visible in Page Layout view
    and on print, but NOT in Normal edit view) — this avoids the AbsoluteAnchor
    floating image intercepting double-click cell-edit events.
    """
    if config.get("pricing_info", {}).get("route_strategy") == "small-segment":
        wb = _generate_xlsx_standard(config)
    else:
        wb = _generate_xlsx_custom(config)
    if config.get("阶梯配置"):
        _xl_add_tiered_sheet(wb, config)
    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    quote_no = config.get('报价编号') or gen_quote_number()
    quote_date = config.get('报价日期') or datetime.now().strftime('%Y年%m月%d日')
    png = _generate_watermark_image(quote_no, quote_date)
    if png:
        from app.domain.render_xlsx_header_watermark import inject_header_watermark
        xlsx_bytes = inject_header_watermark(xlsx_bytes, png)
    return xlsx_bytes
