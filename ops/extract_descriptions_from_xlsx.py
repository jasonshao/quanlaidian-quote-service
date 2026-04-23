#!/usr/bin/env python3
"""Extract per-product 功能说明 text and the 权益类自助充值模块 annotation block
from 全来店底价单V5.xlsx into a plaintext JSON file for the quote renderer.

Pipeline: xlsx → extract (this tool) → references/product_descriptions.json.

Output structure:

    {
      "descriptions": {
        "轻餐|轻餐连锁标准版": "①商户中心：...",
        "正餐|厨房KDS": "使用大屏幕...",
        ...
      },
      "annotations": {
        "权益类自助充值模块": {
          "category": "短信与聚合外卖",
          "text_lines": ["...", "..."]
        }
      }
    }

Keys are "{meal_type}|{name}" so that 轻餐/正餐 with the same product name
(e.g. 厨房KDS) each keep their own description.
"""
import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SHEET_MEAL_TYPE = {
    "轻餐报价单确认": "轻餐",
    "正餐报价单确认": "正餐",
}

# Column indices (1-based) in the 底价单 sheets:
COL_MODULE = 1       # A 模块
COL_CATEGORY = 2     # B 商品分类
COL_NAME = 3         # C 商品名称
COL_UNIT = 4         # D 单位
COL_DESCRIPTION = 8  # H 功能说明 / 套餐说明

# Data rows start at row 6; row 5 is the column header.
DATA_START_ROW = 6

ANNOTATION_MODULE_NAME = "权益类自助充值模块"

# Normalize full-width Chinese parens to half-width so keys line up with
# product_catalog.md (the catalog uses `()` whereas the xlsx uses `（）`).
_PAREN_TRANSLATION = str.maketrans({"（": "(", "）": ")"})


_TRAILING_PARENS_RE = re.compile(r"\s*\([^()]*\)\s*$")

# Circled digits used to delimit sub-module segments in package descriptions.
_CN_NUMBER_MARKERS = "①②③④⑤⑥⑦⑧⑨⑩"
_SEGMENT_SPLIT_RE = re.compile(f"[{_CN_NUMBER_MARKERS}]")
_PACKAGE_HEADER_RE = re.compile(r"^套餐\d+$")
# Section headers in column A (signal end of a package's sub-rows).
_GROUP_KEYWORDS = ("门店优惠套餐", "门店增值模块", "总部模块", "实施服务", "硬件设备")


def split_parent_segments(desc):
    """Split a package's combined description like
    "①商户中心：品牌组织结构…\n②点餐收银：先付后吃模式…"
    into a list of `(prefix, body)` tuples. `prefix` is the label before
    the first `：` (e.g. "商户中心"), `body` is the remaining description.
    Trailing `*…` platform notes are discarded.
    """
    if not desc:
        return []
    parts = _SEGMENT_SPLIT_RE.split(desc)
    segments = []
    for part in parts:
        text = part.strip()
        if not text:
            continue
        # Drop trailing note lines (start with '*' or '＊'), keep pre-note body.
        body_lines = []
        for line in text.splitlines():
            stripped = line.lstrip()
            if stripped.startswith("*") or stripped.startswith("＊"):
                break
            body_lines.append(line)
        text = "\n".join(body_lines).strip()
        if not text:
            continue
        m = re.match(r"^([^：:\n]+)[：:](.+)$", text, re.DOTALL)
        if m:
            prefix = m.group(1).strip()
            body = m.group(2).strip()
        else:
            prefix = ""
            body = text
        # Drop any "（必选）" / "(必选)" etc. suffix from the prefix label.
        prefix_clean = _TRAILING_PARENS_RE.sub("", prefix.translate(_PAREN_TRANSLATION)).strip()
        segments.append((prefix_clean, body))
    return segments


def _match_segment(sub, segments, used):
    """Pick the best (prefix, body) from `segments` for a sub-row.

    Matching preference:
    1. Exact name match     (prefix == sub商品名称 or vice versa — normalized)
    2. Name substring match (prefix in name, or name startswith prefix)
    3. Category match       (prefix == sub商品分类)
    4. Index fallback       — first unused segment if nothing matched
    """
    if not segments:
        return None

    name = normalize_name(sub.get("商品名称") or "")
    name_stripped = strip_trailing_parens(name)
    category = (sub.get("商品分类") or "").strip()

    def score(i, prefix):
        if i in used:
            return -1
        if not prefix:
            return 0
        p = prefix
        if p == name or p == name_stripped:
            return 100
        if name.startswith(p) or name_stripped.startswith(p):
            return 80
        if p.startswith(name) or p.startswith(name_stripped):
            return 70
        if p in name or name in p:
            return 60
        if p == category:
            return 50
        return 0

    best_i, best_score = None, 0
    for i, (prefix, _) in enumerate(segments):
        s = score(i, prefix)
        if s > best_score:
            best_i, best_score = i, s

    if best_i is None:
        # Fallback: first unused segment in order.
        for i in range(len(segments)):
            if i not in used:
                best_i = i
                break
    if best_i is None:
        return None
    used.add(best_i)
    return segments[best_i][1]


def normalize_name(name: str) -> str:
    return name.translate(_PAREN_TRANSLATION).strip()


def strip_trailing_parens(name: str) -> str:
    """Drop a single trailing `(...)` suffix so e.g. `宴秘书标准版套餐(标准版)`
    lines up with the catalog's `宴秘书标准版套餐`."""
    return _TRAILING_PARENS_RE.sub("", name).strip()

# Authoritative text for 权益类自助充值模块 — updated channel unit prices
# (0.039 / 0.032 / 0.021) supersede the older values in the xlsx itself.
ANNOTATION_OVERRIDE = {
    ANNOTATION_MODULE_NAME: {
        "category": "短信与聚合外卖",
        "text_lines": [
            "平台小程序开通后需要按照用量在商户中心后台自行充值手机验证次数，25元500次，150元5000次",
            "三方外卖接单需要按照用量在商户中心后台自行充值通道费：39 元 / 1000 单（单价 0.039元 / 单），159 元 / 5000 单（单价 0.032元 / 单），209 元 / 10000 单（单价 0.021元 / 单）",
        ],
    }
}


def _cell(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    return str(val).strip() or None


def _is_new_group(module_cell):
    if not module_cell:
        return False
    return any(kw in module_cell for kw in _GROUP_KEYWORDS)


def _store_description(descriptions, meal_type, name, desc):
    if not name or not desc:
        return
    normalized = normalize_name(name)
    stripped = strip_trailing_parens(normalized)
    for candidate in (normalized, stripped):
        key = f"{meal_type}|{candidate}"
        if key not in descriptions:
            descriptions[key] = desc


def _store_package_contents(package_contents, meal_type, name, subs):
    if not name or not subs:
        return
    normalized = normalize_name(name)
    stripped = strip_trailing_parens(normalized)
    for candidate in (normalized, stripped):
        key = f"{meal_type}|{candidate}"
        if key not in package_contents:
            package_contents[key] = [dict(s) for s in subs]


def extract(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    descriptions: dict[str, str] = {}
    package_contents: dict[str, list[dict]] = {}

    for sheet_name, meal_type in SHEET_MEAL_TYPE.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        current_package_name = None
        current_parent_desc = None
        current_subs: list[dict] = []

        def flush_package():
            nonlocal current_package_name, current_parent_desc, current_subs
            if current_package_name and current_subs:
                segments = split_parent_segments(current_parent_desc or "")
                used: set[int] = set()
                for sub in current_subs:
                    if sub.get("功能说明"):
                        continue
                    body = _match_segment(sub, segments, used)
                    if body:
                        sub["功能说明"] = body
                _store_package_contents(
                    package_contents, meal_type, current_package_name, current_subs,
                )
            current_package_name = None
            current_parent_desc = None
            current_subs = []

        for row in range(DATA_START_ROW, ws.max_row + 1):
            module = _cell(ws, row, COL_MODULE)
            category = _cell(ws, row, COL_CATEGORY)
            name = _cell(ws, row, COL_NAME)
            unit = _cell(ws, row, COL_UNIT)
            desc = _cell(ws, row, COL_DESCRIPTION)

            # Skip the annotation row — text is handled separately.
            if module == ANNOTATION_MODULE_NAME:
                continue

            # New group (门店增值模块 / 总部模块收费 / 实施服务 / ...) → end the open package.
            if _is_new_group(module):
                flush_package()

            # Package header row: col B like "套餐1" and col C has the package name.
            if category and _PACKAGE_HEADER_RE.match(category) and name:
                flush_package()
                current_package_name = name
                current_parent_desc = desc
                current_subs = []
                _store_description(descriptions, meal_type, name, desc)
                continue

            # Sub-row under an open package: col A empty, col B/C both filled,
            # B is not another package header.
            if (
                current_package_name is not None
                and not module
                and category
                and name
                and not _PACKAGE_HEADER_RE.match(category)
            ):
                current_subs.append({
                    "商品分类": category,
                    "商品名称": name,
                    "单位": unit or "",
                    "功能说明": desc or "",
                })
                _store_description(descriptions, meal_type, name, desc)
                continue

            # Any other row (standalone module / service) — close the package
            # and capture its description as before.
            flush_package()
            _store_description(descriptions, meal_type, name, desc)

        # End of sheet — flush the last package.
        flush_package()

    return {
        "descriptions": descriptions,
        "annotations": dict(ANNOTATION_OVERRIDE),
        "package_contents": package_contents,
    }


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="Source 全来店底价单V5.xlsx path")
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parent.parent / "references" / "product_descriptions.json"),
        help="Output JSON path",
    )
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)

    payload = extract(xlsx_path)
    payload["_meta"] = {
        "version": "v5",
        "source_file": str(xlsx_path),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False),
        encoding="utf-8",
    )
    print(
        f"Wrote descriptions: {output_path} "
        f"({len(payload['descriptions'])} products, "
        f"{len(payload['annotations'])} annotations, "
        f"{len(payload['package_contents'])} packages)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
