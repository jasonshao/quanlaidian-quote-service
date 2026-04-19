#!/usr/bin/env python3
"""
Extract a plaintext pricing baseline JSON from the 全来店底价单V5 xlsx workbook.

Pipeline: xlsx → extract (this tool) → plaintext JSON → obfuscate_baseline.py → .obf.
Ported verbatim from legacy scripts/extract_pricing_baseline_v5.py.
"""
import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SHEET_MEAL_TYPE = {
    "轻餐报价单确认": "轻餐",
    "正餐报价单确认": "正餐",
}

SECTION_KEYWORDS = {
    "门店优惠套餐": "门店套餐",
    "门店增值模块": "门店增值模块",
    "总部模块收费": "总部模块",
    "实施服务收费": "实施服务",
}

VALID_GROUPS = {"门店套餐", "门店增值模块", "总部模块", "实施服务"}


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "赠送"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def detect_group(module_cell, current_group):
    text = str(module_cell or "").strip()
    for keyword, group in SECTION_KEYWORDS.items():
        if keyword in text:
            return group
    return current_group


def extract_items(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    items = {}

    for sheet_name, meal_type in SHEET_MEAL_TYPE.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        current_group = None

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5, values_only=True):
            module, _, name, unit, price = row
            current_group = detect_group(module, current_group)
            if current_group not in VALID_GROUPS:
                continue
            if name is None or str(name).strip() == "":
                continue

            cost_price = parse_number(price)
            if cost_price is None or cost_price <= 0:
                continue

            key = (meal_type, current_group, str(name).strip())
            if key in items:
                continue
            items[key] = {
                "meal_type": meal_type,
                "group": current_group,
                "name": str(name).strip(),
                "unit": str(unit).strip() if unit is not None else "",
                "cost_price": round(cost_price, 4),
            }

    return list(items.values())


def main(argv=None):
    parser = argparse.ArgumentParser(description="Extract pricing baseline from 全来店底价单V5.xlsx")
    parser.add_argument("--xlsx", required=True, help="Source xlsx path")
    parser.add_argument("--output", required=True, help="Output JSON path")
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)

    data = {
        "version": "v5",
        "source_file": str(xlsx_path),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "items": sorted(
            extract_items(xlsx_path),
            key=lambda x: (x["meal_type"], x["group"], x["name"]),
        ),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote baseline: {output_path} ({len(data['items'])} items)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
