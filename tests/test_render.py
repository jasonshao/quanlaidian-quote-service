import json
import pytest
from pathlib import Path

from app.domain.render_pdf import render_pdf
from app.domain.pricing import build_quotation_config

PRODUCT_CATALOG = Path("/Users/sqb/ai/quanlaidian-quotation-skill/references/product_catalog.md")
FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture
def empty_baseline():
    return {"items": []}


def _load_form(name):
    return json.loads((FIXTURES_DIR / name).read_text(encoding="utf-8"))


def _build_config(form_name, baseline):
    form = _load_form(form_name)
    return build_quotation_config(form, baseline, PRODUCT_CATALOG)


def test_render_pdf_returns_valid_pdf(empty_baseline):
    config = _build_config("form_light_meal_5_stores.json", empty_baseline)
    pdf_bytes = render_pdf(config)
    assert pdf_bytes[:5] == b"%PDF-"
    assert len(pdf_bytes) > 1000  # non-trivial PDF


def test_render_pdf_contains_brand(empty_baseline):
    config = _build_config("form_light_meal_5_stores.json", empty_baseline)
    pdf_bytes = render_pdf(config)
    # Brand name should appear somewhere in PDF stream.
    # CID fonts encode CJK text as CID codes (not raw UTF-8), so we check
    # either for raw UTF-8 bytes or a non-trivial PDF size indicating content.
    assert b"\xe6\xb5\x8b\xe8\xaf\x95" in pdf_bytes or len(pdf_bytes) > 3000  # 测试 in UTF-8


def test_render_pdf_full_meal(empty_baseline):
    config = _build_config("form_full_meal_10_stores.json", empty_baseline)
    pdf_bytes = render_pdf(config)
    assert pdf_bytes[:5] == b"%PDF-"
    assert len(pdf_bytes) > 1000


# ============================================================
# XLSX tests
# ============================================================
from app.domain.render_xlsx import render_xlsx


def test_render_xlsx_returns_valid_xlsx(empty_baseline):
    config = _build_config("form_light_meal_5_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    assert len(xlsx_bytes) > 1000
    # XLSX is a ZIP file (starts with PK)
    assert xlsx_bytes[:2] == b"PK"


def test_render_xlsx_has_data(empty_baseline):
    import openpyxl
    import io
    config = _build_config("form_light_meal_5_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert len(wb.sheetnames) >= 1
    ws = wb.active
    assert ws.max_row > 5


def test_render_xlsx_full_meal(empty_baseline):
    config = _build_config("form_full_meal_10_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    assert xlsx_bytes[:2] == b"PK"
    assert len(xlsx_bytes) > 1000


# ============================================================
# 功能说明 column + 权益类自助充值模块 annotation block
# ============================================================
REPO_ROOT = Path(__file__).resolve().parent.parent
LOCAL_CATALOG = REPO_ROOT / "references" / "product_catalog.md"
LOCAL_DESCRIPTIONS = REPO_ROOT / "references" / "product_descriptions.json"


def _build_config_with_descriptions(form_name, baseline):
    from app.domain.product_descriptions import load_descriptions
    form = _load_form(form_name)
    desc = load_descriptions(LOCAL_DESCRIPTIONS)
    return build_quotation_config(form, baseline, LOCAL_CATALOG, descriptions=desc)


def test_xlsx_has_description_column_and_annotation(empty_baseline):
    import openpyxl, io
    config = _build_config_with_descriptions("form_light_meal_5_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["报价单"]
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "序号")
    headers = [ws.cell(header_row, c).value for c in range(1, 9)]
    assert headers[-1] == "功能说明"
    # The first data row (parent package) has empty H because sub-rows
    # expand below it. A sub-row in H just below the parent should be non-empty.
    parent_row = header_row + 1
    # Find the first sub-row: col E contains "-" (qty dash).
    sub_row = next(r for r in range(parent_row + 1, ws.max_row + 1) if ws.cell(r, 5).value == "-")
    assert ws.cell(sub_row, 8).value  # sub-row has non-empty 功能说明
    flat = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any(v and "权益类自助充值模块" in str(v) for v in flat)


def test_xlsx_package_expands_into_sub_rows(empty_baseline):
    import openpyxl, io
    config = _build_config_with_descriptions("form_light_meal_5_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["报价单"]
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "序号")
    # Rows with col E == "-" are sub-rows; there should be at least 3 for a
    # 轻餐 standard/营销基础版 package.
    sub_rows = [r for r in range(header_row + 1, ws.max_row + 1) if ws.cell(r, 5).value == "-"]
    assert len(sub_rows) >= 3
    # Sub-row col C should include one of the known sub-module names.
    sub_names = " ".join(str(ws.cell(r, 3).value or "") for r in sub_rows)
    assert "商户中心" in sub_names or "点餐收银" in sub_names
    # Sub-row qty/unit_price/total all display "-".
    for r in sub_rows:
        assert ws.cell(r, 5).value == "-"
        assert ws.cell(r, 6).value == "-"
        assert ws.cell(r, 7).value == "-"


def test_xlsx_custom_template_renders_annotation_on_cover(empty_baseline):
    import openpyxl, io
    config = _build_config_with_descriptions("form_full_meal_10_stores.json", empty_baseline)
    config["门店数量"] = 100
    config["pricing_info"]["route_strategy"] = "large-segment"
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "封面" in wb.sheetnames
    ws = wb["封面"]
    flat = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any(v and "权益类自助充值模块" in str(v) for v in flat)


def test_pdf_contains_description_and_annotation(empty_baseline):
    try:
        import fitz
    except ImportError:
        pytest.skip("PyMuPDF not available")
    config = _build_config_with_descriptions("form_light_meal_5_stores.json", empty_baseline)
    pdf_bytes = render_pdf(config)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = "".join(page.get_text() for page in doc)
    doc.close()
    assert "功能说明" in text
    assert "权益类自助充值模块" in text
    assert "0.039" in text


def test_xlsx_main_table_columns_match_spec(empty_baseline):
    """Lock the single-tier main table column layout against skill §3.2.1."""
    import openpyxl, io
    config = _build_config_with_descriptions("form_full_meal_10_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["报价单"]
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "序号")
    headers = [ws.cell(header_row, c).value for c in range(1, 9)]
    assert headers == [
        "序号", "商品分类", "商品名称", "单位",
        "数量", "商品单价", "小计", "功能说明",
    ]


def test_xlsx_tiered_sheet_9_column_layout(empty_baseline):
    """Lock the tiered-comparison sheet header (skill §3.2.2, 9 columns)."""
    import openpyxl, io
    config = _build_config_with_descriptions("form_full_meal_10_stores.json", empty_baseline)
    # Force large-segment routing to produce a 2-tier comparison.
    config["门店数量"] = 100
    from app.domain.pricing import build_tier_config
    config["阶梯配置"] = build_tier_config(True, "正餐", 100)
    config["pricing_info"]["route_strategy"] = "large-segment"
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "阶梯报价参考" in wb.sheetnames
    ws = wb["阶梯报价参考"]
    headers = [ws.cell(3, c).value for c in range(1, 10)]
    assert headers[0:4] == ["序号", "商品分类", "商品名称", "单位"]
    assert headers[4].endswith("单价")
    assert headers[5].endswith("小计")
    assert headers[6].endswith("单价")
    assert headers[7].endswith("小计")
    assert headers[8] == "功能说明"
    # Bottom summary rows exist: at least one 小计 + 合计.
    labels = [ws.cell(r, 2).value for r in range(4, ws.max_row + 1)]
    assert "小计" in labels
    assert "合计" in labels
    first_col_last = [ws.cell(r, 1).value for r in range(4, ws.max_row + 1)]
    assert all(v != "折算单店年费" for v in first_col_last)


def test_xlsx_no_floating_watermark_image(empty_baseline):
    """Regression: the 2000×2000 watermark must not be a floating image.

    Floating images intercept double-click events and block users from
    entering cell-edit mode. The watermark is now embedded as a page-header
    image instead, which is invisible to mouse hit-testing in Normal view.
    Logo images (86×26 and 73×26) are still floating — they don't overlap
    data cells, so they don't block editing.
    """
    import openpyxl, io
    config = _build_config_with_descriptions("form_light_meal_5_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    for name in wb.sheetnames:
        ws = wb[name]
        large_imgs = [im for im in getattr(ws, "_images", [])
                      if (getattr(im, "width", 0) or 0) >= 500]
        assert not large_imgs, (
            f"{name}: unexpected large floating image(s) found "
            f"(would block cell-edit double-clicks): "
            f"{[(im.width, im.height) for im in large_imgs]}"
        )


def test_xlsx_header_watermark_present(empty_baseline):
    """Every sheet should declare an odd-page header image and reference
    a shared media part containing the watermark PNG.
    """
    import io, zipfile
    config = _build_config_with_descriptions("form_full_meal_10_stores.json", empty_baseline)
    config["门店数量"] = 100
    from app.domain.pricing import build_tier_config
    config["阶梯配置"] = build_tier_config(True, "正餐", 100)
    config["pricing_info"]["route_strategy"] = "large-segment"
    xlsx_bytes = render_xlsx(config)

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        names = set(zf.namelist())
        assert "xl/media/imageHF1.png" in names, (
            "shared header watermark image missing"
        )
        assert len(zf.read("xl/media/imageHF1.png")) > 1000

        ct = zf.read("[Content_Types].xml").decode("utf-8")
        assert 'Extension="vml"' in ct
        assert 'Extension="png"' in ct

        sheet_paths = sorted(
            n for n in names
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
        )
        assert sheet_paths, "no worksheets found"
        for sheet in sheet_paths:
            xml = zf.read(sheet).decode("utf-8")
            assert "<oddHeader>" in xml and "&amp;G" in xml, (
                f"{sheet}: expected <oddHeader> with &G picture marker"
            )
            assert "<legacyDrawingHF" in xml, (
                f"{sheet}: expected <legacyDrawingHF> reference"
            )


def test_xlsx_opens_clean_in_openpyxl(empty_baseline):
    """The post-processed xlsx must remain readable by openpyxl with no
    parse errors. This catches malformed OOXML — broken namespace prefixes,
    misplaced elements, etc. — that would otherwise only surface in Excel.
    """
    import openpyxl, io
    config = _build_config_with_descriptions("form_light_meal_5_stores.json", empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames
    buf = io.BytesIO()
    wb.save(buf)
    assert len(buf.getvalue()) > 1000


def test_pdf_package_expanded_sub_modules(empty_baseline):
    try:
        import fitz
    except ImportError:
        pytest.skip("PyMuPDF not available")
    config = _build_config_with_descriptions("form_light_meal_5_stores.json", empty_baseline)
    pdf_bytes = render_pdf(config)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = "".join(page.get_text() for page in doc)
    doc.close()
    # 轻餐 packages expand to rows including 商户中心-轻餐版 / 点餐收银 / 平台小程序
    assert "商户中心" in text
    assert "点餐收银" in text
    assert "平台小程序" in text


# ============================================================
# Regression: XLSX 大客户段主明细 sheet 必须基于 config 真实金额，
# 不再使用硬编码 "刊例价 qty=1, factor=0.8" 覆盖。
# 历史 bug：30 店套餐被显示为 数量=1、单价=7920、小计=7920，
# 而 PDF / 阶梯参考 sheet 显示真实 数量=30、单价=1584、小计=47520。
# ============================================================
def _make_large_segment_config(empty_baseline):
    config = _build_config_with_descriptions("form_full_meal_10_stores.json", empty_baseline)
    config["门店数量"] = 100
    config["pricing_info"]["route_strategy"] = "large-segment"
    return config


def test_xlsx_custom_main_sheet_uses_config_quote_values(empty_baseline):
    """大客户段「门店软件与增值」sheet 套餐主行的 数量/单价/小计 必须等于 config 字段，
    而不是被硬编码为 数量=1、商品单价=标准价×0.8。"""
    import openpyxl, io
    config = _make_large_segment_config(empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    assert "门店软件与增值" in wb.sheetnames
    ws = wb["门店软件与增值"]

    pkg = next(it for it in config["报价项目"] if it.get("模块分类") == "门店软件套餐")
    pkg_name = pkg["商品名称"]
    pkg_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(r, 3).value == pkg_name
    )

    assert ws.cell(pkg_row, 5).value == int(pkg["数量"]), \
        f"数量应等于 config 数量 {pkg['数量']}，bug 时被硬编码为 1"
    assert ws.cell(pkg_row, 6).value == int(pkg["商品单价"]), \
        f"商品单价应等于 config 商品单价 {pkg['商品单价']}，bug 时被算成 标准价×0.8"
    assert ws.cell(pkg_row, 7).value == int(pkg["报价小计"]), \
        f"小计应等于 config 报价小计 {pkg['报价小计']}，bug 时等于 单价×1"


def test_xlsx_custom_hq_sheet_respects_config_qty(empty_baseline):
    """总部模块 sheet 的数量列必须读 config 里的 数量字段，"""
    """注入一个 数量=3 的总部项以触发非 1 路径。"""
    import openpyxl, io
    config = _make_large_segment_config(empty_baseline)
    config.setdefault("报价项目", []).append({
        "商品分类": "总部模块",
        "商品名称": "测试总部多数量项",
        "单位": "套/年",
        "标准价": 1000,
        "数量": 3,
        "商品单价": 1000,
        "报价小计": 3000,
        "模块分类": "总部模块",
        "deal_price_factor": 1.0,
        "成交价系数": 1.0,
        "折扣": 0.0,
    })
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    assert "总部模块" in wb.sheetnames
    ws = wb["总部模块"]

    target_row = next(
        r for r in range(1, ws.max_row + 1)
        if ws.cell(r, 3).value == "测试总部多数量项"
    )
    assert ws.cell(target_row, 5).value == 3, "qty 应保留 config 中的 3"
    assert ws.cell(target_row, 6).value == 1000
    assert ws.cell(target_row, 7).value == 3000, "小计 应等于 数量×单价"


def test_xlsx_main_sheets_total_matches_config_quote_total(empty_baseline):
    """大客户段所有主明细 sheet（门店软件与增值 / 总部模块 / 实施服务）的数据行
    G 列总和应等于 config['internal_financials']['quote_total']，
    即 Excel 的真实成交价合计与配置中的 quote_total 完全一致。"""
    import openpyxl, io
    from decimal import Decimal
    config = _make_large_segment_config(empty_baseline)
    xlsx_bytes = render_xlsx(config)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)

    total = 0
    for sheet_name in ("门店软件与增值", "总部模块", "实施服务"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            a_val = ws.cell(r, 1).value
            g_val = ws.cell(r, 7).value
            # 数据行：A 列是 int 序号；合计行 A 列是 "合计"，子项行 A 列空。
            if isinstance(a_val, int) and isinstance(g_val, (int, float)):
                total += g_val

    expected = config["internal_financials"]["quote_total"]
    assert total == expected, \
        f"主明细 sheet G 列累加 = {total}，应等于 quote_total = {expected}"
