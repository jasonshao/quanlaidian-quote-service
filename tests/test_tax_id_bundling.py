"""Regression tests for the 电子发票接口 → 电子发票-税号 auto-append rule (issue #8).

When the request's 门店增值模块 contains 电子发票接口, the quotation MUST include an
extra row 电子发票-税号 with 单位=税号/年 and 数量=form["税号数量"] (defaulting to 1).
When 电子发票接口 is absent, the row MUST NOT appear.
"""
import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.domain.pricing import build_quotation_config
from app.domain.product_descriptions import load_descriptions
from app.domain.render_xlsx import render_xlsx

REPO_ROOT = Path(__file__).resolve().parent.parent
LOCAL_CATALOG = REPO_ROOT / "references" / "product_catalog.md"
LOCAL_DESCRIPTIONS = REPO_ROOT / "references" / "product_descriptions.json"


@pytest.fixture
def tax_id_baseline():
    """Baseline with 电子发票-税号 / 电子发票接口 cost for both meal types."""
    return {
        "items": [
            {"meal_type": "轻餐", "group": "门店增值模块", "name": "电子发票接口", "unit": "店/年", "cost_price": 365.0},
            {"meal_type": "轻餐", "group": "门店增值模块", "name": "电子发票-税号", "unit": "税号/年", "cost_price": 209.0},
            {"meal_type": "正餐", "group": "门店增值模块", "name": "电子发票接口", "unit": "店/年", "cost_price": 365.0},
            {"meal_type": "正餐", "group": "门店增值模块", "name": "电子发票-税号", "unit": "税号/年", "cost_price": 209.0},
        ]
    }


def _form(**overrides):
    base = {
        "客户品牌名称": "测试",
        "餐饮类型": "正餐",
        "门店数量": 10,
        "门店套餐": "正餐连锁营销旗舰版",
        "门店增值模块": [],
        "总部模块": [],
        "配送中心数量": 0,
        "生产加工中心数量": 0,
        "是否启用阶梯报价": False,
    }
    base.update(overrides)
    return base


def _config(form, baseline):
    desc = load_descriptions(LOCAL_DESCRIPTIONS)
    return build_quotation_config(form, baseline, LOCAL_CATALOG, descriptions=desc)


def _tax_id_rows(config):
    return [it for it in config["报价项目"] if it.get("商品名称") == "电子发票-税号"]


def test_tax_id_row_appended_with_explicit_qty(tax_id_baseline):
    cfg = _config(_form(门店增值模块=["电子发票接口"], 税号数量=3), tax_id_baseline)
    rows = _tax_id_rows(cfg)
    assert len(rows) == 1
    row = rows[0]
    assert row["数量"] == 3
    assert row["单位"] == "税号/年"
    assert row["商品分类"] == "增值模块"
    assert row["模块分类"] == "门店增值模块"
    # cost 209 × 1.20 = 250.8 → 251
    assert row["商品单价"] == 251
    assert row["报价小计"] == 251 * 3


def test_tax_id_row_defaults_to_qty_one(tax_id_baseline):
    cfg = _config(_form(门店增值模块=["电子发票接口"]), tax_id_baseline)
    rows = _tax_id_rows(cfg)
    assert len(rows) == 1
    assert rows[0]["数量"] == 1


def test_tax_id_row_absent_without_invoice_interface(tax_id_baseline):
    cfg = _config(_form(门店增值模块=[], 税号数量=5), tax_id_baseline)
    assert _tax_id_rows(cfg) == []


def test_tax_id_row_absent_with_other_modules_only(tax_id_baseline):
    # Baseline also needs 成本管理 to avoid fallback quirks; use catalog fallback
    baseline = dict(tax_id_baseline)
    cfg = _config(_form(门店增值模块=["成本管理"]), baseline)
    assert _tax_id_rows(cfg) == []


def test_tax_id_row_appears_in_xlsx_main_table(tax_id_baseline):
    cfg = _config(_form(门店增值模块=["电子发票接口"], 税号数量=2), tax_id_baseline)
    xlsx = render_xlsx(cfg)
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["报价单"]
    header_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "序号")
    tax_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, 3).value == "电子发票-税号":
            tax_row = r
            break
    assert tax_row is not None, "电子发票-税号 row missing from 报价单 sheet"
    assert ws.cell(tax_row, 4).value == "税号/年"
    assert ws.cell(tax_row, 5).value == 2
    assert ws.cell(tax_row, 6).value == 251  # 商品单价
    assert ws.cell(tax_row, 7).value == 502  # 小计


def test_schema_accepts_tax_id_qty():
    from app.domain.schema import QuoteForm
    form = QuoteForm(
        客户品牌名称="A", 餐饮类型="正餐", 门店数量=10, 门店套餐="正餐连锁营销旗舰版",
        门店增值模块=["电子发票接口"], 税号数量=4,
    )
    assert form.税号数量 == 4


def test_schema_defaults_tax_id_qty_to_one():
    from app.domain.schema import QuoteForm
    form = QuoteForm(
        客户品牌名称="A", 餐饮类型="正餐", 门店数量=10, 门店套餐="正餐连锁营销旗舰版",
    )
    assert form.税号数量 == 1


def test_schema_rejects_zero_tax_id_qty():
    from app.domain.schema import QuoteForm
    from pydantic import ValidationError
    with pytest.raises(ValidationError):
        QuoteForm(
            客户品牌名称="A", 餐饮类型="正餐", 门店数量=10, 门店套餐="正餐连锁营销旗舰版",
            税号数量=0,
        )
